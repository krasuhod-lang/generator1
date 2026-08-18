"""Слой 1 — Очередь URL (персистентная, возобновляемая).

Очередь хранится в SQLite: каждой строке присваивается уникальный ``task_id``,
статус (``pending``/``in_progress``/``success``/``failed``/``blocked``) и
счётчик попыток ``retry_count``. Благодаря персистентности процесс можно
продолжить с места остановки (отключение сервера, OOM, рестарт), а не начинать
заново — задачи в статусе ``in_progress`` при старте возвращаются в ``pending``.

Ингест источника (CSV / Excel / Google Sheets) выполняется идемпотентно:
повторная загрузка того же URL не создаёт дубликат.
"""

from __future__ import annotations

import csv
import io
import ipaddress
import os
import socket
import sqlite3
import urllib.request
from contextlib import closing
from typing import Iterable, Iterator, List, Optional
from urllib.parse import urlparse

from .config import CONFIG
from .models import Task, TaskStatus

_SCHEMA = """
CREATE TABLE IF NOT EXISTS url_queue (
    task_id     INTEGER PRIMARY KEY AUTOINCREMENT,
    url         TEXT NOT NULL UNIQUE,
    status      TEXT NOT NULL DEFAULT 'pending',
    retry_count INTEGER NOT NULL DEFAULT 0,
    error       TEXT,
    result      TEXT,
    created_at  TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
    updated_at  TIMESTAMP DEFAULT CURRENT_TIMESTAMP
);
CREATE INDEX IF NOT EXISTS idx_url_queue_status ON url_queue(status);
"""


class UrlQueue:
    """Персистентная очередь URL поверх SQLite.

    Пример::

        q = UrlQueue("queue.sqlite3")
        q.ingest_iterable(["https://a.example", "https://b.example"])
        task = q.next_pending()
        ...
        q.mark_success(task.task_id, result={...})
    """

    def __init__(self, db_path: str, max_retries: int = 3, ingest_dir: Optional[str] = None) -> None:
        self.db_path = db_path
        self.max_retries = max_retries
        # Каталог, из которого разрешено читать файлы для ингеста (Слой 1).
        # Ограничивает path traversal: файлы вне него отклоняются.
        base = ingest_dir if ingest_dir is not None else CONFIG.get("ingest_dir", "")
        self.ingest_dir = os.path.realpath(base or os.getcwd())
        # ``check_same_thread=False`` — очередь может использоваться воркерами.
        self._conn = sqlite3.connect(db_path, check_same_thread=False)
        self._conn.row_factory = sqlite3.Row
        self._conn.executescript(_SCHEMA)
        self._conn.commit()

    # -- ingest ---------------------------------------------------------------
    def ingest_iterable(self, urls: Iterable[str]) -> int:
        """Добавить URL из произвольного итерабельного источника.

        Возвращает количество реально добавленных (новых) URL. Пустые строки и
        дубликаты игнорируются (идемпотентность).
        """
        added = 0
        with self._conn:  # транзакция
            for raw in urls:
                url = (raw or "").strip()
                if not url:
                    continue
                cur = self._conn.execute(
                    "INSERT OR IGNORE INTO url_queue(url) VALUES (?)", (url,)
                )
                added += cur.rowcount
        return added

    def ingest_csv(self, path: str, url_column: str = "url") -> int:
        """Ингест из CSV-файла с колонкой ``url_column``."""
        safe_path = self._safe_ingest_path(path)
        with open(safe_path, "r", encoding="utf-8-sig", newline="") as fh:
            return self._ingest_csv_stream(fh, url_column)

    def ingest_google_sheet(self, sheet_url: str, url_column: str = "url") -> int:
        """Ингест из Google Sheets.

        Принимает как обычную ссылку на таблицу, так и прямой CSV-export URL —
        ссылка нормализуется в ``.../export?format=csv``. URL ограничен
        доверенными доменами Google по http(s) (защита от SSRF).
        """
        export_url = _google_sheet_csv_url(sheet_url)
        # Инлайн-барьер против SSRF: разрешаем только http(s) на доверенные
        # домены Google; дополнительно проверяем, что хост не резолвится в
        # приватную сеть.
        parsed = urlparse(export_url)
        host = (parsed.hostname or "").lower()
        if parsed.scheme not in ("http", "https") or host not in _ALLOWED_SHEET_HOSTS:
            raise ValueError(f"Недопустимый URL Google Sheets: {export_url!r}")
        _assert_public_host(host, parsed.port or 443)
        req = urllib.request.Request(export_url, method="GET")
        with urllib.request.urlopen(req, timeout=30) as resp:  # noqa: S310 (URL валидирован)
            data = resp.read().decode("utf-8-sig")
        return self._ingest_csv_stream(io.StringIO(data), url_column)

    def ingest_excel(self, path: str, url_column: str = "url") -> int:
        """Ингест из Excel (.xlsx). Требует пакет ``openpyxl``."""
        try:
            from openpyxl import load_workbook
        except ImportError as exc:  # pragma: no cover - зависит от окружения
            raise RuntimeError(
                "Для ингеста Excel установите зависимость 'openpyxl'"
            ) from exc

        wb = load_workbook(self._safe_ingest_path(path), read_only=True, data_only=True)
        try:
            ws = wb.active
            rows = ws.iter_rows(values_only=True)
            try:
                header = [str(c).strip() if c is not None else "" for c in next(rows)]
            except StopIteration:
                return 0
            try:
                col_idx = header.index(url_column)
            except ValueError:
                raise KeyError(
                    f"Колонка '{url_column}' не найдена в Excel: {header}"
                )
            urls = (
                str(row[col_idx]).strip()
                for row in rows
                if col_idx < len(row) and row[col_idx] not in (None, "")
            )
            return self.ingest_iterable(urls)
        finally:
            wb.close()

    def _safe_ingest_path(self, path: str) -> str:
        """Разрешить путь к файлу ингеста только внутри ``self.ingest_dir``.

        Защита от path traversal: используется только базовое имя файла
        (все компоненты каталога отбрасываются), после чего путь нормализуется
        и обязан находиться внутри разрешённого каталога.
        """
        # Отбрасываем любые компоненты каталога из недоверенного ввода —
        # это устраняет обход вида ``../../etc/passwd``.
        name = os.path.basename(path)
        if not name or name in (".", ".."):
            raise ValueError(f"Недопустимое имя файла ингеста: {path!r}")
        candidate = os.path.realpath(os.path.join(self.ingest_dir, name))
        if os.path.commonpath([self.ingest_dir, candidate]) != self.ingest_dir:
            raise ValueError(
                f"Путь ингеста вне разрешённого каталога {self.ingest_dir!r}: {path!r}"
            )
        return candidate

    def _ingest_csv_stream(self, fh, url_column: str) -> int:
        reader = csv.DictReader(fh)
        if reader.fieldnames is None or url_column not in reader.fieldnames:
            raise KeyError(
                f"Колонка '{url_column}' не найдена в источнике: {reader.fieldnames}"
            )
        return self.ingest_iterable(row.get(url_column, "") for row in reader)

    # -- consume --------------------------------------------------------------
    def next_pending(self) -> Optional[Task]:
        """Атомарно взять следующую ожидающую задачу и пометить in_progress.

        Возвращает ``None``, если ожидающих задач нет.
        """
        with self._conn:  # транзакция гарантирует, что задачу возьмёт один воркер
            row = self._conn.execute(
                "SELECT * FROM url_queue WHERE status = ? "
                "ORDER BY task_id LIMIT 1",
                (TaskStatus.PENDING.value,),
            ).fetchone()
            if row is None:
                return None
            self._conn.execute(
                "UPDATE url_queue SET status = ?, updated_at = CURRENT_TIMESTAMP "
                "WHERE task_id = ?",
                (TaskStatus.IN_PROGRESS.value, row["task_id"]),
            )
        return Task(
            task_id=row["task_id"],
            url=row["url"],
            status=TaskStatus.IN_PROGRESS,
            retry_count=row["retry_count"],
            error=row["error"],
        )

    def recover_stale(self) -> int:
        """Вернуть «зависшие» in_progress задачи в pending (после рестарта).

        Возвращает число восстановленных задач.
        """
        with self._conn:
            cur = self._conn.execute(
                "UPDATE url_queue SET status = ?, updated_at = CURRENT_TIMESTAMP "
                "WHERE status = ?",
                (TaskStatus.PENDING.value, TaskStatus.IN_PROGRESS.value),
            )
        return cur.rowcount

    # -- state transitions ----------------------------------------------------
    def mark_success(self, task_id: int, result: Optional[str] = None) -> None:
        self._set_status(task_id, TaskStatus.SUCCESS, result=result, error=None)

    def mark_blocked(self, task_id: int, error: Optional[str] = None) -> None:
        self._set_status(task_id, TaskStatus.BLOCKED, error=error)

    def mark_failure(self, task_id: int, error: Optional[str] = None) -> TaskStatus:
        """Учесть неуспех: инкремент retry_count.

        Если попытки не исчерпаны — задача возвращается в ``pending`` для
        повтора (Слой 6). Иначе переводится в терминальный ``failed``.
        Возвращает итоговый статус.
        """
        with self._conn:
            row = self._conn.execute(
                "SELECT retry_count FROM url_queue WHERE task_id = ?", (task_id,)
            ).fetchone()
            if row is None:
                raise KeyError(f"Задача {task_id} не найдена")
            retry_count = row["retry_count"] + 1
            if retry_count >= self.max_retries:
                new_status = TaskStatus.FAILED
            else:
                new_status = TaskStatus.PENDING
            self._conn.execute(
                "UPDATE url_queue SET status = ?, retry_count = ?, error = ?, "
                "updated_at = CURRENT_TIMESTAMP WHERE task_id = ?",
                (new_status.value, retry_count, error, task_id),
            )
        return new_status

    def _set_status(
        self,
        task_id: int,
        status: TaskStatus,
        result: Optional[str] = None,
        error: Optional[str] = None,
    ) -> None:
        with self._conn:
            self._conn.execute(
                "UPDATE url_queue SET status = ?, result = ?, error = ?, "
                "updated_at = CURRENT_TIMESTAMP WHERE task_id = ?",
                (status.value, result, error, task_id),
            )

    # -- introspection --------------------------------------------------------
    def get(self, task_id: int) -> Optional[Task]:
        row = self._conn.execute(
            "SELECT * FROM url_queue WHERE task_id = ?", (task_id,)
        ).fetchone()
        if row is None:
            return None
        return Task(
            task_id=row["task_id"],
            url=row["url"],
            status=TaskStatus(row["status"]),
            retry_count=row["retry_count"],
            error=row["error"],
            result=row["result"],
        )

    def counts(self) -> dict:
        """Сводка по статусам — для Слоя 8 (мониторинг)."""
        rows = self._conn.execute(
            "SELECT status, COUNT(*) AS n FROM url_queue GROUP BY status"
        ).fetchall()
        return {r["status"]: r["n"] for r in rows}

    def pending_count(self) -> int:
        row = self._conn.execute(
            "SELECT COUNT(*) AS n FROM url_queue WHERE status = ?",
            (TaskStatus.PENDING.value,),
        ).fetchone()
        return int(row["n"])

    def iter_pending(self) -> Iterator[Task]:
        """Итерировать по задачам, пока в очереди есть ожидающие."""
        while True:
            task = self.next_pending()
            if task is None:
                return
            yield task

    def all_tasks(self) -> List[Task]:
        rows = self._conn.execute(
            "SELECT * FROM url_queue ORDER BY task_id"
        ).fetchall()
        return [
            Task(
                task_id=r["task_id"],
                url=r["url"],
                status=TaskStatus(r["status"]),
                retry_count=r["retry_count"],
                error=r["error"],
                result=r["result"],
            )
            for r in rows
        ]

    def close(self) -> None:
        with closing(self._conn):
            pass

    def __enter__(self) -> "UrlQueue":
        return self

    def __exit__(self, *exc) -> None:
        self.close()


def _google_sheet_csv_url(sheet_url: str) -> str:
    """Нормализовать ссылку Google Sheets в CSV-export URL."""
    if "format=csv" in sheet_url:
        return sheet_url
    if "/edit" in sheet_url:
        base = sheet_url.split("/edit", 1)[0]
        gid = ""
        if "gid=" in sheet_url:
            gid = "&gid=" + sheet_url.split("gid=", 1)[1].split("&", 1)[0]
        return f"{base}/export?format=csv{gid}"
    return sheet_url


# Доверенные хосты Google для ингеста таблиц (защита от SSRF).
_ALLOWED_SHEET_HOSTS = frozenset(
    {"docs.google.com", "sheets.googleapis.com", "drive.google.com"}
)


def _assert_public_host(host: str, port: int) -> None:
    """Отклонить хост, резолвящийся в приватный/loopback адрес (anti-SSRF)."""
    try:
        infos = socket.getaddrinfo(host, port, proto=socket.IPPROTO_TCP)
    except OSError as exc:
        raise ValueError(f"Не удалось разрешить хост {host!r}: {exc}") from exc
    for info in infos:
        ip = ipaddress.ip_address(info[4][0])
        if ip.is_private or ip.is_loopback or ip.is_link_local or ip.is_reserved:
            raise ValueError(f"Хост {host!r} резолвится в приватный адрес {ip}")


def _validate_google_sheet_url(url: str) -> None:
    """Проверить, что URL безопасен для загрузки: http(s) + домен Google.

    Используется в тестах; основная проверка встроена в ``ingest_google_sheet``.
    """
    parsed = urlparse(url)
    host = (parsed.hostname or "").lower()
    if parsed.scheme not in {"http", "https"}:
        raise ValueError(f"Недопустимая схема URL: {parsed.scheme!r}")
    if host not in _ALLOWED_SHEET_HOSTS:
        raise ValueError(f"Хост не в списке доверенных Google-доменов: {host!r}")
    _assert_public_host(host, parsed.port or 443)


__all__ = ["UrlQueue"]
